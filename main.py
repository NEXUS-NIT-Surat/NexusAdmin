import os
import openpyxl
import requests
from dotenv import load_dotenv
load_dotenv()

if __name__ == "__main__":
    try:
        # Get the Token from the server
        clAdminJSON = requests.post(os.getenv('BACKEND_URL')+"api/user/login",data={
            "email":os.getenv('ADMIN_EMAIL_ID'),
            "password":os.getenv('ADMIN_PASSWORD')
        })
        clAdminData = clAdminJSON.json()

        # Get the Forms list from the server
        clFormsList = requests.get(os.getenv('BACKEND_URL')+"forms")
        clForms = clFormsList.json()
        for index,form in enumerate(clForms):
            print(index+1,form["name"])
        sFormID = int(input("Enter the Form Index:"))-1
        
        # Get the Form Responses of the selected Form ID
        headers = {
            'Authorization': f"Bearer {clAdminData['token']}",
            'Content-Type': 'application/json'
        }
        clFormsJSON = requests.get(os.getenv('BACKEND_URL')+"forms/get-responses/"+clForms[sFormID]['_id'],headers=headers)
        clFormsResponses = clFormsJSON.json()
        clFormsResponses = clFormsResponses['responses']

        # Fill the details in the Excel Sheet
        sExcelFilePath = './NexusResponses.xlsx'
        workbook = openpyxl.Workbook()
        if not os.path.exists(sExcelFilePath):
            workbook.save(sExcelFilePath)

        workbook = openpyxl.load_workbook(sExcelFilePath)
        existing_sheet = workbook[clForms[sFormID]['name']] if clForms[sFormID]['name'] in workbook.sheetnames else None
        if existing_sheet is None:
            clSheet = workbook.create_sheet(clForms[sFormID]['name'])
        clSheet = workbook[clForms[sFormID]['name']]
        clColumNames = list(clFormsResponses[0].keys())

        # Setting up the Column Fields for the Excel File
        for iColumnIndexInExcelFile, sEachColumnName in enumerate(clColumNames, 1):
            clSheet.cell(row=1, column=iColumnIndexInExcelFile, value=sEachColumnName)

        # Filling up the Data in the Excel File
        for iRowNumberInExcelFile, clData in enumerate(clFormsResponses, 2):
            for iColumnNumberInExcelFile, sColumnName in enumerate(clColumNames, 1):
                clSheet.cell(row=iRowNumberInExcelFile, column=iColumnNumberInExcelFile, value=clData[sColumnName])

        workbook.save(sExcelFilePath)
    except Exception as e:
        print("Found Error:",str(e))